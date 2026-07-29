"""
Import students from Student.xlsx into the `students` table.

Source file layout: one sheet per department, each sheet containing repeated
attendance-form blocks (one block per ระดับชั้น/ชั้นปีที่/ห้องที่/ภาคเรียน).
Each block looks like:

    ...แผนกวิชา...ระดับชั้น...ชั้นปีที่...ห้องที่...
    ภาคเรียนที่...ปีการศึกษา...ครูที่ปรึกษา...
    (boilerplate row)
    ลำดับที่ | รหัสประจำตัว | ชื่อ - สกุล | | | ลายมือชื่อ | หมายเหตุ
    1 | <id> | <prefix> | <first_name> | <last_name> | | ...
    ...

Some students appear in multiple blocks across terms (re-recorded roster each
semester). We keep only the row with the highest (academic_year, semester),
since the check-in system needs each student's current department/year/room.
"""

import re

import openpyxl

from db import get_db_connection

EXCEL_PATH = "Student.xlsx"

DEPT_RE = re.compile(
    r"แผนกวิชา(?P<department>.*?)ระดับชั้น(?P<level>.*?)ชั้นปีที่(?P<year_level>.*?)ห้องที่(?P<room>.*)"
)
TERM_RE = re.compile(
    r"ภาคเรียนที่(?P<semester>.*?)ปีการศึกษา(?P<academic_year>.*?)ครูที่ปรึกษา(?P<advisor>.*)"
)


def clean(value):
    """Strip the dot/ellipsis fill-in-the-blank characters used on the paper form."""
    if value is None:
        return ""
    text = re.sub(r"[.…]{2,}", " ", str(value))
    return text.strip(" .…")


def extract_students(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        n = len(rows)
        i = 0
        while i < n:
            cell = rows[i][0]
            if not cell or "แผนกวิชา" not in str(cell):
                i += 1
                continue

            dept_match = DEPT_RE.search(str(cell))
            if not dept_match:
                i += 1
                continue
            info = {k: clean(v) for k, v in dept_match.groupdict().items()}

            semester = academic_year = ""
            term_cell = rows[i + 1][0] if i + 1 < n else None
            if term_cell:
                term_match = TERM_RE.search(str(term_cell))
                if term_match:
                    term_info = {k: clean(v) for k, v in term_match.groupdict().items()}
                    semester, academic_year = term_info["semester"], term_info["academic_year"]

            header_row = None
            for offset in range(5):
                idx = i + 1 + offset
                if idx < n and rows[idx][0] and str(rows[idx][0]).strip() == "ลำดับที่":
                    header_row = idx
                    break
            if header_row is None:
                i += 1
                continue

            r = header_row + 1
            while r < n:
                seq = rows[r][0]
                try:
                    int(seq)
                except (TypeError, ValueError):
                    break
                student_id = str(rows[r][1]).strip()
                records.append(
                    {
                        "student_id": student_id,
                        "prefix": clean(rows[r][2]),
                        "first_name": clean(rows[r][3]),
                        "last_name": clean(rows[r][4]),
                        "department": info["department"],
                        "level": info["level"],
                        "year_level": info["year_level"],
                        "room": info["room"],
                        "semester": semester,
                        "academic_year": academic_year,
                    }
                )
                r += 1
            i = r

    return records


def dedupe_keep_latest_term(records):
    latest = {}
    for rec in records:
        term_key = (int(rec["academic_year"] or 0), int(rec["semester"] or 0))
        existing = latest.get(rec["student_id"])
        if existing is None or term_key > existing[0]:
            latest[rec["student_id"]] = (term_key, rec)
    return [rec for _, rec in latest.values()]


def import_to_mysql(records):
    conn = get_db_connection()
    cur = conn.cursor()
    sql = """
        INSERT INTO students
            (student_id, prefix, first_name, last_name, department, level, year_level, room, semester, academic_year)
        VALUES
            (%(student_id)s, %(prefix)s, %(first_name)s, %(last_name)s, %(department)s, %(level)s, %(year_level)s, %(room)s, %(semester)s, %(academic_year)s)
        ON DUPLICATE KEY UPDATE
            prefix=VALUES(prefix), first_name=VALUES(first_name), last_name=VALUES(last_name),
            department=VALUES(department), level=VALUES(level), year_level=VALUES(year_level),
            room=VALUES(room), semester=VALUES(semester), academic_year=VALUES(academic_year)
    """
    cur.executemany(sql, records)
    conn.commit()
    cur.close()
    conn.close()


if __name__ == "__main__":
    all_records = extract_students(EXCEL_PATH)
    print(f"Parsed {len(all_records)} rows from Excel")

    deduped = dedupe_keep_latest_term(all_records)
    print(f"{len(deduped)} unique students after keeping latest term per student_id")

    import_to_mysql(deduped)
    print("Import complete")
