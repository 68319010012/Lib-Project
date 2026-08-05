import calendar
import csv
import io
import os
import re
from datetime import datetime, timedelta
from functools import wraps

import bcrypt
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import Workbook
from openpyxl.styles import Font

from db import get_db_connection

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key-change-before-production")
# NOTE: dev-only default. Set FLASK_SECRET_KEY in .env with a real random value before deploying.

limiter = Limiter(get_remote_address, app=app, storage_uri="memory://")
# NOTE: in-memory storage resets on restart and isn't shared across worker
# processes — fine for the dev server; switch to Redis storage_uri before
# running multiple workers in production.


@app.errorhandler(429)
def rate_limit_exceeded(_error):
    return jsonify(error="เข้าสู่ระบบผิดพลาดหลายครั้งเกินไป กรุณารอสักครู่แล้วลองใหม่"), 429

def gender_from_prefix(prefix):
    prefix = (prefix or "").strip()
    if prefix == "นาย":
        return "ชาย"
    if prefix in ("นาง", "นางสาว"):
        return "หญิง"
    return "ไม่ระบุ"


def _aggregate_checkin_period(cur, start_date, end_date):
    """Aggregate checkin_logs rows in [start_date, end_date] inclusive (YYYY-MM-DD
    strings). Shared by every dashboard-style report so period-over-period
    comparisons all use the same counting logic."""
    cur.execute(
        """
        SELECT s.student_id, s.department, s.level, s.year_level, c.timestamp
        FROM checkin_logs c
        JOIN users u ON u.user_id = c.user_id
        JOIN students s ON s.student_id = u.student_id
        WHERE DATE(c.timestamp) BETWEEN %s AND %s
        """,
        (start_date, end_date),
    )
    rows = cur.fetchall()

    day_counts = {}
    dept_counts = {}
    level_counts = {}
    for row in rows:
        day_key = row["timestamp"].strftime("%Y-%m-%d")
        day_counts[day_key] = day_counts.get(day_key, 0) + 1
        dept = row["department"] or "ไม่ระบุแผนก"
        dept_counts[dept] = dept_counts.get(dept, 0) + 1
        level_label = f"{row['level']}.{row['year_level']}" if row["level"] and row["year_level"] else "ไม่ระบุ"
        level_counts[level_label] = level_counts.get(level_label, 0) + 1

    days_with_data = len(day_counts)
    total_events = len(rows)
    busiest_day = None
    if day_counts:
        busiest_key, busiest_count = max(day_counts.items(), key=lambda item: item[1])
        busiest_day = {"day": busiest_key, "count": busiest_count}

    return {
        "total_events": total_events,
        "unique_students": len({row["student_id"] for row in rows}),
        "avg_daily": round(total_events / days_with_data) if days_with_data else 0,
        "busiest_day": busiest_day,
        "day_counts": day_counts,
        "dept_counts": dept_counts,
        "level_counts": level_counts,
    }


def _pct_delta(current, previous):
    """% change from previous -> current, or None if there's no previous-period
    baseline to compare against (avoids a division by zero reading as 0%)."""
    if not previous:
        return None
    return round((current - previous) / previous * 100, 1)


def _month_bounds(month_str):
    """'YYYY-MM' -> (first_day, last_day) as 'YYYY-MM-DD' strings."""
    year, mon = (int(part) for part in month_str.split("-"))
    last_day = calendar.monthrange(year, mon)[1]
    return f"{month_str}-01", f"{month_str}-{last_day:02d}"


def _previous_month(month_str):
    year, mon = (int(part) for part in month_str.split("-"))
    if mon == 1:
        return f"{year - 1}-12"
    return f"{year}-{mon - 1:02d}"


def _ranked_breakdown(counts, limit=8):
    """dict of label->count -> sorted list of {name, count, pct} vs the top entry,
    for rendering as horizontal bar rows."""
    top = sorted(counts.items(), key=lambda item: -item[1])[:limit]
    max_count = top[0][1] if top else 0
    return [
        {"name": name, "count": count, "pct": round((count / max_count) * 100) if max_count else 0}
        for name, count in top
    ]


def _export_response(filename_base, sections, fmt):
    """Build a CSV/Excel download from one or more (title, headers, data_rows)
    sections. Every export route (daily/monthly/department/dashboard/executive/
    compare) funnels through this so the two file formats stay consistent."""
    if fmt == "csv":
        buf = io.StringIO()
        buf.write("﻿")  # BOM so Excel opens Thai text as UTF-8 instead of mis-decoding it
        writer = csv.writer(buf)
        for i, (title, headers, data_rows) in enumerate(sections):
            if i > 0:
                writer.writerow([])
            if len(sections) > 1:
                writer.writerow([title])
            writer.writerow(headers)
            writer.writerows(data_rows)
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    if fmt == "xlsx":
        wb = Workbook()
        wb.remove(wb.active)
        for title, headers, data_rows in sections:
            ws = wb.create_sheet(title=title[:31])  # Excel sheet-name limit
            ws.append(headers)
            for row in data_rows:
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{filename_base}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return jsonify(error="รูปแบบไฟล์ไม่ถูกต้อง"), 400


LIBRARY_CLOSING_TIME = os.environ.get("LIBRARY_CLOSING_TIME", "17:00")
# NOTE: single closing time for every day — doesn't yet account for weekends/holidays
# having different hours. Extend _closing_datetime() if that's ever needed.
OVERDUE_HOURS = 6
TIME_RE = re.compile(r"([01]\d|2[0-3]):[0-5]\d")


def _closing_datetime(reference):
    """`reference`'s date combined with LIBRARY_CLOSING_TIME, as a datetime."""
    hour, minute = (int(part) for part in LIBRARY_CLOSING_TIME.split(":"))
    return reference.replace(hour=hour, minute=minute, second=0, microsecond=0)


def compute_planned_checkout(duration_minutes, checkout_time, now=None):
    """Returns the planned checkout datetime, clamped to today's closing time —
    or None, meaning "until closing" (no specific time chosen)."""
    now = now or datetime.now()
    closing = _closing_datetime(now)

    if duration_minutes is not None:
        planned = now + timedelta(minutes=duration_minutes)
    elif checkout_time:
        hour, minute = (int(part) for part in checkout_time.split(":"))
        planned = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    else:
        return None

    return min(planned, closing)


def get_current_checkins(cur):
    """Latest checkin_logs row per user, for users whose latest row is type='in'
    (i.e. everyone currently checked in). `cur` must be a dictionary cursor."""
    cur.execute(
        """
        SELECT c.log_id, c.user_id, c.timestamp, c.planned_checkout_at
        FROM checkin_logs c
        INNER JOIN (
            SELECT user_id, MAX(log_id) AS max_log_id
            FROM checkin_logs
            GROUP BY user_id
        ) latest ON c.user_id = latest.user_id AND c.log_id = latest.max_log_id
        WHERE c.type = 'in'
        """
    )
    return cur.fetchall()


def auto_checkout_job():
    """Runs every 2 minutes (see scheduler setup at the bottom of this file):
    force-checks-out anyone past their planned_checkout_at."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        now = datetime.now()
        due = [
            row
            for row in get_current_checkins(cur)
            if row["planned_checkout_at"] and row["planned_checkout_at"] <= now
        ]
        for row in due:
            cur.execute(
                "INSERT INTO checkin_logs (user_id, type, checkout_source) VALUES (%s, 'out', 'auto')",
                (row["user_id"],),
            )
        if due:
            conn.commit()
    finally:
        cur.close()
        conn.close()


def retire_expired_accounts():
    """Runs daily (see scheduler setup at the bottom of this file): suspends any
    student account whose users.created_at is a full academic year (365 days) in
    the past. Admin accounts are never retired."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE users
            SET account_status = 'retired'
            WHERE role = 'student' AND account_status = 'approved'
              AND created_at <= NOW() - INTERVAL 1 YEAR
            """
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return jsonify(error="กรุณาเข้าสู่ระบบ"), 401
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify(error="ต้องใช้สิทธิ์แอดมิน"), 403
        return view(*args, **kwargs)

    return wrapped


def page_login_required(view):
    """Like login_required, but redirects to /login instead of returning JSON 401 —
    for page routes navigated to directly by the browser, not fetch()."""

    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login_page"))
        return view(*args, **kwargs)

    return wrapped


def page_admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if session.get("role") != "admin":
            return redirect(url_for("dashboard_page"))
        return view(*args, **kwargs)

    return wrapped


@app.route("/")
def index():
    return redirect(url_for("login_page"))


@app.route("/login")
def login_page():
    return render_template("pages/login.html")


@app.route("/signup")
def signup_page():
    return render_template("pages/signup.html")


@app.route("/dashboard")
@page_login_required
def dashboard_page():
    return render_template("pages/dashboard.html")


@app.route("/admin/dashboard")
@page_login_required
@page_admin_required
def admin_dashboard_page():
    return render_template("pages/admin_dashboard.html")


@app.route("/admin/members-management")
@page_login_required
@page_admin_required
def members_management_page():
    return render_template("pages/members_management.html")


@app.route("/admin/attendance-logs")
@page_login_required
@page_admin_required
def attendance_logs_page():
    return render_template("pages/attendance_logs.html")


@app.route("/profile")
@page_login_required
def edit_profile_page():
    return render_template("pages/edit_profile.html")


@app.route("/register", methods=["POST"])
def register():
    data = request.get_json(silent=True) or request.form
    student_id = (data.get("student_id") or "").strip()
    username = student_id
    password = data.get("password") or ""
    prefix = (data.get("prefix") or "").strip()
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    department = (data.get("department") or "").strip()
    level = (data.get("level") or "").strip()
    year_level = (data.get("year_level") or "").strip()

    if not student_id or not password:
        return jsonify(error="กรุณากรอกรหัสนักศึกษาและรหัสผ่าน"), 400
    if not prefix or not first_name or not last_name or not department:
        return jsonify(error="กรุณากรอกคำนำหน้า ชื่อ นามสกุล และแผนกวิชา"), 400
    if level not in ("ปวช", "ปวส"):
        return jsonify(error="ระดับชั้นต้องเป็น ปวช หรือ ปวส"), 400
    valid_years = ("1", "2", "3") if level == "ปวช" else ("1", "2")
    if year_level not in valid_years:
        return jsonify(error=f"ชั้นปีของ {level} ต้องเป็นหนึ่งใน {', '.join(valid_years)}"), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT user_id FROM users WHERE student_id = %s", (student_id,))
        if cur.fetchone() is not None:
            return jsonify(error="นักศึกษาคนนี้มีบัญชีอยู่แล้ว"), 409

        # Student profile is entered manually at signup — upsert rather than require
        # the student_id to already exist in the imported roster.
        cur.execute(
            """
            INSERT INTO students (student_id, prefix, first_name, last_name, department, level, year_level)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                prefix = VALUES(prefix), first_name = VALUES(first_name), last_name = VALUES(last_name),
                department = VALUES(department), level = VALUES(level), year_level = VALUES(year_level)
            """,
            (student_id, prefix, first_name, last_name, department, level, year_level),
        )

        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

        cur.execute(
            """
            INSERT INTO users (username, password_hash, role, student_id, account_status)
            VALUES (%s, %s, 'student', %s, 'approved')
            """,
            (username, password_hash, student_id),
        )
        conn.commit()
        user_id = cur.lastrowid
    finally:
        cur.close()
        conn.close()

    session["user_id"] = user_id
    session["role"] = "student"
    session["student_id"] = student_id
    return jsonify(message="สร้างบัญชีสำเร็จ", role="student"), 201


@app.route("/login", methods=["POST"])
@limiter.limit("5 per minute")
def login():
    data = request.get_json(silent=True) or request.form
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not username or not password:
        return jsonify(error="กรุณากรอกชื่อผู้ใช้และรหัสผ่าน"), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if user is None or not bcrypt.checkpw(password.encode("utf-8"), user["password_hash"].encode("utf-8")):
        return jsonify(error="ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง"), 401
    if user["account_status"] == "retired":
        return jsonify(error="บัญชีนี้ถูกระงับเนื่องจากครบกำหนด 1 ปีการศึกษา กรุณาติดต่อฝ่ายทะเบียน"), 403
    if user["account_status"] != "approved":
        return jsonify(error="บัญชียังไม่ได้รับการอนุมัติจากแอดมิน"), 403

    session["user_id"] = user["user_id"]
    session["role"] = user["role"]
    session["student_id"] = user["student_id"]
    return jsonify(message="เข้าสู่ระบบสำเร็จ", role=user["role"])


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify(message="ออกจากระบบแล้ว")


@app.route("/checkin", methods=["POST"])
@login_required
def checkin():
    user_id = session["user_id"]
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT type FROM checkin_logs WHERE user_id = %s ORDER BY log_id DESC LIMIT 1",
            (user_id,),
        )
        last = cur.fetchone()
        next_type = "out" if last and last["type"] == "in" else "in"
        planned_checkout_at = None

        if next_type == "in":
            duration_minutes = data.get("duration_minutes")
            checkout_time = (data.get("checkout_time") or "").strip() or None

            if duration_minutes is not None:
                try:
                    duration_minutes = int(duration_minutes)
                except (TypeError, ValueError):
                    return jsonify(error="duration_minutes ต้องเป็นตัวเลข"), 400
                if duration_minutes <= 0:
                    return jsonify(error="duration_minutes ต้องมากกว่า 0"), 400

            if checkout_time is not None and not TIME_RE.fullmatch(checkout_time):
                return jsonify(error="รูปแบบ checkout_time ต้องเป็น HH:MM"), 400

            planned_checkout_at = compute_planned_checkout(duration_minutes, checkout_time)
            if planned_checkout_at is not None and planned_checkout_at <= datetime.now():
                return jsonify(error="เวลาที่เลือกต้องอยู่ในอนาคต"), 400

            cur.execute(
                "INSERT INTO checkin_logs (user_id, type, planned_checkout_at) VALUES (%s, 'in', %s)",
                (user_id, planned_checkout_at),
            )
        else:
            cur.execute(
                "INSERT INTO checkin_logs (user_id, type, checkout_source) VALUES (%s, 'out', 'manual')",
                (user_id,),
            )

        conn.commit()
        checkin_message = "เช็คอินสำเร็จ" if next_type == "in" else "เช็คเอาต์สำเร็จ"
        return jsonify(
            message=checkin_message,
            type=next_type,
            planned_checkout_at=planned_checkout_at.isoformat() if planned_checkout_at else None,
        )
    finally:
        cur.close()
        conn.close()


@app.route("/checkin/extend", methods=["POST"])
@login_required
def checkin_extend():
    user_id = session["user_id"]
    data = request.get_json(silent=True) or request.form
    try:
        extend_minutes = int(data.get("minutes"))
    except (TypeError, ValueError):
        return jsonify(error="minutes ต้องเป็นตัวเลข"), 400
    if extend_minutes <= 0:
        return jsonify(error="minutes ต้องมากกว่า 0"), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT log_id, type, planned_checkout_at FROM checkin_logs WHERE user_id = %s ORDER BY log_id DESC LIMIT 1",
            (user_id,),
        )
        last = cur.fetchone()
        if last is None or last["type"] != "in":
            return jsonify(error="ไม่มีสถานะเช็คอินค้างอยู่"), 400
        if last["planned_checkout_at"] is None:
            return jsonify(error="เลือกจนกว่าจะปิดไว้ ไม่มีเวลาให้ต่อ"), 400

        new_planned = min(
            last["planned_checkout_at"] + timedelta(minutes=extend_minutes),
            _closing_datetime(datetime.now()),
        )
        cur.execute(
            "UPDATE checkin_logs SET planned_checkout_at = %s WHERE log_id = %s",
            (new_planned, last["log_id"]),
        )
        conn.commit()
        return jsonify(message="ต่อเวลาสำเร็จ", planned_checkout_at=new_planned.isoformat())
    finally:
        cur.close()
        conn.close()


@app.route("/library-info", methods=["GET"])
@login_required
def library_info():
    return jsonify(closing_time=LIBRARY_CLOSING_TIME)


@app.route("/me", methods=["GET"])
@login_required
def me():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT u.user_id, u.username, u.role, u.account_status,
                   s.student_id, s.prefix, s.first_name, s.last_name,
                   s.department, s.level, s.year_level, s.room
            FROM users u
            LEFT JOIN students s ON s.student_id = u.student_id
            WHERE u.user_id = %s
            """,
            (session["user_id"],),
        )
        return jsonify(cur.fetchone())
    finally:
        cur.close()
        conn.close()


@app.route("/profile/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json(silent=True) or request.form
    current_password = data.get("current_password") or ""
    new_password = data.get("new_password") or ""

    if not current_password or not new_password:
        return jsonify(error="กรุณากรอกรหัสผ่านปัจจุบันและรหัสผ่านใหม่"), 400
    if len(new_password) < 8:
        return jsonify(error="รหัสผ่านใหม่ต้องมีอย่างน้อย 8 ตัวอักษร"), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT password_hash FROM users WHERE user_id = %s", (session["user_id"],))
        user = cur.fetchone()
        if user is None or not bcrypt.checkpw(
            current_password.encode("utf-8"), user["password_hash"].encode("utf-8")
        ):
            return jsonify(error="รหัสผ่านปัจจุบันไม่ถูกต้อง"), 401

        new_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        cur.execute("UPDATE users SET password_hash = %s WHERE user_id = %s", (new_hash, session["user_id"]))
        conn.commit()
        return jsonify(message="เปลี่ยนรหัสผ่านสำเร็จ")
    finally:
        cur.close()
        conn.close()


@app.route("/me/history", methods=["GET"])
@login_required
def my_history():
    try:
        limit = int(request.args.get("limit", 20))
    except ValueError:
        limit = 20
    limit = max(1, min(limit, 100))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT type, timestamp, planned_checkout_at FROM checkin_logs WHERE user_id = %s ORDER BY log_id DESC LIMIT %s",
            (session["user_id"], limit),
        )
        rows = cur.fetchall()
        for row in rows:
            row["timestamp"] = row["timestamp"].isoformat()
            row["planned_checkout_at"] = row["planned_checkout_at"].isoformat() if row["planned_checkout_at"] else None
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/admin/members", methods=["GET"])
@login_required
@admin_required
def admin_members():
    search = request.args.get("search", "").strip()
    department = request.args.get("department", "").strip()
    level = request.args.get("level", "").strip()
    year_level = request.args.get("year_level", "").strip()

    conditions = ["u.account_status = 'approved'"]
    params = []
    if search:
        conditions.append(
            "(s.first_name LIKE %s OR s.last_name LIKE %s OR s.student_id LIKE %s OR u.username LIKE %s)"
        )
        like = f"%{search}%"
        params.extend([like, like, like, like])
    if department:
        conditions.append("s.department = %s")
        params.append(department)
    if level:
        conditions.append("s.level = %s")
        params.append(level)
    if year_level:
        conditions.append("s.year_level = %s")
        params.append(year_level)

    where_clause = " AND ".join(conditions)

    sql = f"""
        SELECT u.user_id, u.username, s.student_id, s.prefix, s.first_name, s.last_name,
               s.department, s.level, s.year_level, s.room,
               (SELECT MAX(c.timestamp) FROM checkin_logs c WHERE c.user_id = u.user_id) AS last_visit
        FROM users u
        JOIN students s ON s.student_id = u.student_id
        WHERE {where_clause}
        ORDER BY s.first_name, s.last_name
    """

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()
        for row in rows:
            row["last_visit"] = row["last_visit"].isoformat() if row["last_visit"] else None
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/admin/checkins/current", methods=["GET"])
@login_required
@admin_required
def admin_checkins_current():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        current_rows = get_current_checkins(cur)
        if not current_rows:
            return jsonify([])

        user_ids = [row["user_id"] for row in current_rows]
        placeholders = ",".join(["%s"] * len(user_ids))
        cur.execute(
            f"""
            SELECT u.user_id, s.student_id, s.prefix, s.first_name, s.last_name, s.department, s.level, s.year_level
            FROM users u
            JOIN students s ON s.student_id = u.student_id
            WHERE u.user_id IN ({placeholders})
            """,
            user_ids,
        )
        student_by_user = {row["user_id"]: row for row in cur.fetchall()}
    finally:
        cur.close()
        conn.close()

    now = datetime.now()
    results = []
    for row in current_rows:
        student = student_by_user.get(row["user_id"])
        if student is None:
            continue
        duration_minutes = int((now - row["timestamp"]).total_seconds() // 60)
        results.append(
            {
                **student,
                "checked_in_at": row["timestamp"].isoformat(),
                "planned_checkout_at": row["planned_checkout_at"].isoformat() if row["planned_checkout_at"] else None,
                "duration_minutes": duration_minutes,
                "is_overdue": duration_minutes >= OVERDUE_HOURS * 60,
            }
        )
    results.sort(key=lambda r: r["duration_minutes"], reverse=True)
    return jsonify(results)


@app.route("/admin/checkins/force-checkout", methods=["POST"])
@login_required
@admin_required
def admin_force_checkout():
    data = request.get_json(silent=True) or request.form
    try:
        target_user_id = int(data.get("user_id"))
    except (TypeError, ValueError):
        return jsonify(error="user_id ไม่ถูกต้อง"), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT type FROM checkin_logs WHERE user_id = %s ORDER BY log_id DESC LIMIT 1",
            (target_user_id,),
        )
        last = cur.fetchone()
        if last is None or last["type"] != "in":
            return jsonify(error="ผู้ใช้นี้ไม่ได้ค้างสถานะเช็คอินอยู่"), 400

        cur.execute(
            "INSERT INTO checkin_logs (user_id, type, checkout_source) VALUES (%s, 'out', 'admin_forced')",
            (target_user_id,),
        )
        conn.commit()
        return jsonify(message="บังคับเช็คเอาท์สำเร็จ")
    finally:
        cur.close()
        conn.close()


@app.route("/announcement", methods=["GET"])
@login_required
def get_announcement():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT message, updated_at FROM announcements WHERE id = 1")
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if row is None or not row["message"]:
        return jsonify(message=None, updated_at=None)
    return jsonify(message=row["message"], updated_at=row["updated_at"].isoformat())


@app.route("/admin/announcement", methods=["POST"])
@login_required
@admin_required
def set_announcement():
    data = request.get_json(silent=True) or request.form
    message = (data.get("message") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO announcements (id, message, updated_by)
            VALUES (1, %s, %s)
            ON DUPLICATE KEY UPDATE message = VALUES(message), updated_by = VALUES(updated_by)
            """,
            (message or None, session["user_id"]),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify(message="บันทึกประกาศสำเร็จ")


@app.route("/admin/reports", methods=["GET"])
@login_required
@admin_required
def admin_reports():
    date = request.args.get("date")
    month = request.args.get("month")
    academic_year = request.args.get("academic_year")

    conditions = []
    params = []
    if date:
        conditions.append("DATE(c.timestamp) = %s")
        params.append(date)
    if month:
        conditions.append("DATE_FORMAT(c.timestamp, '%Y-%m') = %s")
        params.append(month)
    if academic_year:
        conditions.append("s.academic_year = %s")
        params.append(academic_year)

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    sql = f"""
        SELECT s.student_id, s.prefix, s.first_name, s.last_name, s.department, s.level, s.year_level,
               c.type, c.timestamp
        FROM checkin_logs c
        JOIN users u ON u.user_id = c.user_id
        JOIN students s ON s.student_id = u.student_id
        {where_clause}
        ORDER BY c.timestamp DESC
    """

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()
        for row in rows:
            row["timestamp"] = row["timestamp"].isoformat()
        return jsonify(rows)
    finally:
        cur.close()
        conn.close()


@app.route("/admin/reports/print", methods=["GET"])
@page_admin_required
def admin_reports_select():
    today = datetime.now().strftime("%Y-%m-%d")
    this_month = datetime.now().strftime("%Y-%m")
    return render_template(
        "reports_select.html", today=today, this_month=this_month, prev_month=_previous_month(this_month)
    )


@app.route("/admin/reports/print/daily", methods=["GET"])
@page_admin_required
def admin_report_daily():
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")

    sql = """
        SELECT s.student_id, s.prefix, s.first_name, s.last_name, s.department, s.level, s.year_level,
               c.type, c.timestamp
        FROM checkin_logs c
        JOIN users u ON u.user_id = c.user_id
        JOIN students s ON s.student_id = u.student_id
        WHERE DATE(c.timestamp) = %s
        ORDER BY s.student_id, c.timestamp
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, (date,))
        logs = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    by_student = {}
    for log in logs:
        row = by_student.setdefault(
            log["student_id"],
            {
                "student_id": log["student_id"],
                "prefix": log["prefix"],
                "first_name": log["first_name"],
                "last_name": log["last_name"],
                "gender": gender_from_prefix(log["prefix"]),
                "department": log["department"],
                "level": log["level"],
                "year_level": log["year_level"],
                "time_in": None,
                "time_out": None,
            },
        )
        time_str = log["timestamp"].strftime("%H:%M:%S")
        if log["type"] == "in" and row["time_in"] is None:
            row["time_in"] = time_str
        if log["type"] == "out":
            row["time_out"] = time_str

    rows = sorted(by_student.values(), key=lambda r: r["student_id"])

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        headers = ["ลำดับ", "รหัสนักศึกษา", "ชื่อ-สกุล", "เพศ", "แผนกวิชา", "ระดับชั้น/ปี", "เวลาเข้า", "เวลาออก"]
        data_rows = [
            [
                i + 1, r["student_id"], f"{r['prefix']}{r['first_name']} {r['last_name']}", r["gender"],
                r["department"], f"{r['level']} ปีที่ {r['year_level']}", r["time_in"] or "-", r["time_out"] or "-",
            ]
            for i, r in enumerate(rows)
        ]
        return _export_response(f"รายงานประจำวัน_{date}", [("รายงานประจำวัน", headers, data_rows)], fmt)

    return render_template(
        "report_daily.html", date=date, rows=rows,
        csv_url=url_for("admin_report_daily", date=date, format="csv"),
        xlsx_url=url_for("admin_report_daily", date=date, format="xlsx"),
    )


@app.route("/admin/reports/print/monthly", methods=["GET"])
@page_admin_required
def admin_report_monthly():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")

    sql = """
        SELECT s.student_id, s.prefix, s.first_name, s.last_name, s.department, s.level, s.year_level,
               COUNT(CASE WHEN c.type = 'in' THEN 1 END) AS checkin_count,
               MAX(c.timestamp) AS last_checkin
        FROM checkin_logs c
        JOIN users u ON u.user_id = c.user_id
        JOIN students s ON s.student_id = u.student_id
        WHERE DATE_FORMAT(c.timestamp, '%Y-%m') = %s
        GROUP BY s.student_id, s.prefix, s.first_name, s.last_name, s.department, s.level, s.year_level
        ORDER BY s.student_id
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, (month,))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    for row in rows:
        row["gender"] = gender_from_prefix(row["prefix"])
        row["last_checkin"] = row["last_checkin"].strftime("%Y-%m-%d %H:%M:%S") if row["last_checkin"] else "-"

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        headers = ["ลำดับ", "รหัสนักศึกษา", "ชื่อ-สกุล", "เพศ", "แผนกวิชา", "ระดับชั้น/ปี", "จำนวนครั้งที่เช็คอิน", "เช็คอินล่าสุด"]
        data_rows = [
            [
                i + 1, r["student_id"], f"{r['prefix']}{r['first_name']} {r['last_name']}", r["gender"],
                r["department"], f"{r['level']} ปีที่ {r['year_level']}", r["checkin_count"], r["last_checkin"],
            ]
            for i, r in enumerate(rows)
        ]
        return _export_response(f"รายงานสรุปรายเดือน_{month}", [("รายงานสรุปรายเดือน", headers, data_rows)], fmt)

    return render_template(
        "report_monthly.html", month=month, rows=rows,
        csv_url=url_for("admin_report_monthly", month=month, format="csv"),
        xlsx_url=url_for("admin_report_monthly", month=month, format="xlsx"),
    )


@app.route("/admin/reports/print/department", methods=["GET"])
@page_admin_required
def admin_report_department():
    academic_year = request.args.get("academic_year")

    conditions = []
    params = []
    if academic_year:
        conditions.append("s.academic_year = %s")
        params.append(academic_year)
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    sql = f"""
        SELECT s.department,
               COUNT(DISTINCT s.student_id) AS student_count,
               COUNT(c.log_id) AS checkin_count
        FROM checkin_logs c
        JOIN users u ON u.user_id = c.user_id
        JOIN students s ON s.student_id = u.student_id
        {where_clause}
        GROUP BY s.department
        ORDER BY s.department
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        headers = ["ลำดับ", "แผนกวิชา", "จำนวนนักศึกษาที่เช็คชื่อ (ไม่ซ้ำคน)", "จำนวนครั้งที่เช็คอินรวม"]
        data_rows = [
            [i + 1, r["department"], r["student_count"], r["checkin_count"]] for i, r in enumerate(rows)
        ]
        name = f"รายงานสรุปตามแผนกวิชา_{academic_year}" if academic_year else "รายงานสรุปตามแผนกวิชา"
        return _export_response(name, [("รายงานสรุปตามแผนกวิชา", headers, data_rows)], fmt)

    export_args = {"academic_year": academic_year} if academic_year else {}
    return render_template(
        "report_department.html", academic_year=academic_year, rows=rows,
        csv_url=url_for("admin_report_department", format="csv", **export_args),
        xlsx_url=url_for("admin_report_department", format="xlsx", **export_args),
    )


@app.route("/admin/reports/print/dashboard", methods=["GET"])
@page_admin_required
def admin_report_dashboard():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    start_date, end_date = _month_bounds(month)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        agg = _aggregate_checkin_period(cur, start_date, end_date)
    finally:
        cur.close()
        conn.close()

    year, mon = (int(part) for part in month.split("-"))
    days_in_month = calendar.monthrange(year, mon)[1]
    daily_trend = [
        {"day": f"{d:02d}", "count": agg["day_counts"].get(f"{month}-{d:02d}", 0)}
        for d in range(1, days_in_month + 1)
    ]
    max_daily = max((d["count"] for d in daily_trend), default=0)
    for d in daily_trend:
        d["pct"] = round((d["count"] / max_daily) * 100) if max_daily else 0

    busiest_day = None
    if agg["busiest_day"]:
        busiest_date = datetime.strptime(agg["busiest_day"]["day"], "%Y-%m-%d")
        busiest_day = {"day": busiest_date.strftime("%d"), "count": agg["busiest_day"]["count"]}

    dept_breakdown = _ranked_breakdown(agg["dept_counts"])
    level_breakdown = _ranked_breakdown(agg["level_counts"])

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        sections = [
            (
                "KPI สรุป", ["ตัวชี้วัด", "ค่า"],
                [
                    ["จำนวนรายการทั้งหมด", agg["total_events"]],
                    ["นักศึกษาไม่ซ้ำคน", agg["unique_students"]],
                    ["เฉลี่ยต่อวัน", agg["avg_daily"]],
                    ["วันที่มีคนใช้มากที่สุด", f"{busiest_day['day']} ({busiest_day['count']} รายการ)" if busiest_day else "-"],
                ],
            ),
            ("แนวโน้มรายวัน", ["วันที่", "จำนวนรายการ"], [[d["day"], d["count"]] for d in daily_trend]),
            ("สัดส่วนตามแผนก", ["แผนกวิชา", "จำนวนรายการ"], [[d["name"], d["count"]] for d in dept_breakdown]),
            ("สัดส่วนตามระดับชั้น", ["ระดับ/ชั้นปี", "จำนวนรายการ"], [[d["name"], d["count"]] for d in level_breakdown]),
        ]
        return _export_response(f"แดชบอร์ดภาพรวม_{month}", sections, fmt)

    return render_template(
        "report_dashboard.html",
        month=month,
        total_events=agg["total_events"],
        unique_students=agg["unique_students"],
        avg_daily=agg["avg_daily"],
        busiest_day=busiest_day,
        daily_trend=daily_trend,
        dept_breakdown=dept_breakdown,
        level_breakdown=level_breakdown,
        csv_url=url_for("admin_report_dashboard", month=month, format="csv"),
        xlsx_url=url_for("admin_report_dashboard", month=month, format="xlsx"),
    )


@app.route("/admin/reports/print/executive", methods=["GET"])
@page_admin_required
def admin_report_executive():
    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    start_date, end_date = _month_bounds(month)
    prev_start, prev_end = _month_bounds(_previous_month(month))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        agg = _aggregate_checkin_period(cur, start_date, end_date)
        prev_agg = _aggregate_checkin_period(cur, prev_start, prev_end)
    finally:
        cur.close()
        conn.close()

    top_depts = _ranked_breakdown(agg["dept_counts"], limit=3)
    busiest_day = None
    if agg["busiest_day"]:
        busiest_date = datetime.strptime(agg["busiest_day"]["day"], "%Y-%m-%d")
        busiest_day = {"day": busiest_date.strftime("%d/%m"), "count": agg["busiest_day"]["count"]}

    total_delta = _pct_delta(agg["total_events"], prev_agg["total_events"])
    unique_delta = _pct_delta(agg["unique_students"], prev_agg["unique_students"])

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        sections = [
            (
                "สรุปสำหรับผู้บริหาร", ["ตัวชี้วัด", "ค่า", "% เทียบเดือนก่อน"],
                [
                    ["จำนวนรายการทั้งหมด", agg["total_events"], total_delta if total_delta is not None else "-"],
                    ["นักศึกษาที่มาใช้บริการ", agg["unique_students"], unique_delta if unique_delta is not None else "-"],
                    ["เฉลี่ยต่อวัน", agg["avg_daily"], "-"],
                    ["วันที่มีคนใช้มากที่สุด", f"{busiest_day['day']} ({busiest_day['count']} รายการ)" if busiest_day else "-", "-"],
                ],
            ),
            ("แผนกที่เข้าใช้มากที่สุด", ["อันดับ", "แผนกวิชา", "จำนวนรายการ"], [[i + 1, d["name"], d["count"]] for i, d in enumerate(top_depts)]),
        ]
        return _export_response(f"สรุปสำหรับผู้บริหาร_{month}", sections, fmt)

    return render_template(
        "report_executive.html",
        month=month,
        total_events=agg["total_events"],
        unique_students=agg["unique_students"],
        avg_daily=agg["avg_daily"],
        busiest_day=busiest_day,
        top_depts=top_depts,
        total_delta=total_delta,
        unique_delta=unique_delta,
        csv_url=url_for("admin_report_executive", month=month, format="csv"),
        xlsx_url=url_for("admin_report_executive", month=month, format="xlsx"),
    )


@app.route("/admin/reports/print/compare", methods=["GET"])
@page_admin_required
def admin_report_compare():
    this_month = datetime.now().strftime("%Y-%m")
    month_a = request.args.get("month_a") or _previous_month(this_month)
    month_b = request.args.get("month_b") or this_month

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        start_a, end_a = _month_bounds(month_a)
        start_b, end_b = _month_bounds(month_b)
        agg_a = _aggregate_checkin_period(cur, start_a, end_a)
        agg_b = _aggregate_checkin_period(cur, start_b, end_b)
    finally:
        cur.close()
        conn.close()

    metrics = [
        {
            "label": "จำนวนรายการทั้งหมด",
            "a": agg_a["total_events"],
            "b": agg_b["total_events"],
            "delta": _pct_delta(agg_b["total_events"], agg_a["total_events"]),
        },
        {
            "label": "นักศึกษาไม่ซ้ำคน",
            "a": agg_a["unique_students"],
            "b": agg_b["unique_students"],
            "delta": _pct_delta(agg_b["unique_students"], agg_a["unique_students"]),
        },
        {
            "label": "เฉลี่ยต่อวัน",
            "a": agg_a["avg_daily"],
            "b": agg_b["avg_daily"],
            "delta": _pct_delta(agg_b["avg_daily"], agg_a["avg_daily"]),
        },
    ]

    dept_names = sorted(set(agg_a["dept_counts"]) | set(agg_b["dept_counts"]))
    dept_compare = sorted(
        (
            {"name": name, "a": agg_a["dept_counts"].get(name, 0), "b": agg_b["dept_counts"].get(name, 0)}
            for name in dept_names
        ),
        key=lambda row: -(row["a"] + row["b"]),
    )[:8]
    max_dept_compare = max((max(row["a"], row["b"]) for row in dept_compare), default=0)
    for row in dept_compare:
        row["pct_a"] = round((row["a"] / max_dept_compare) * 100) if max_dept_compare else 0
        row["pct_b"] = round((row["b"] / max_dept_compare) * 100) if max_dept_compare else 0

    fmt = request.args.get("format")
    if fmt in ("csv", "xlsx"):
        sections = [
            (
                "เปรียบเทียบตัวชี้วัด", ["ตัวชี้วัด", month_a, month_b, "% ผลต่าง"],
                [[m["label"], m["a"], m["b"], m["delta"] if m["delta"] is not None else "-"] for m in metrics],
            ),
            (
                "เปรียบเทียบตามแผนกวิชา", ["แผนกวิชา", month_a, month_b],
                [[d["name"], d["a"], d["b"]] for d in dept_compare],
            ),
        ]
        return _export_response(f"เปรียบเทียบ_{month_a}_กับ_{month_b}", sections, fmt)

    return render_template(
        "report_compare.html",
        month_a=month_a,
        month_b=month_b,
        metrics=metrics,
        dept_compare=dept_compare,
        csv_url=url_for("admin_report_compare", month_a=month_a, month_b=month_b, format="csv"),
        xlsx_url=url_for("admin_report_compare", month_a=month_a, month_b=month_b, format="xlsx"),
    )


# FLASK_DEBUG controls Werkzeug's debug/reload behavior for `python app.py` (dev).
# Defaults to on so local dev keeps auto-reloading exactly like before this
# variable existed; production (.env / Coolify) must set FLASK_DEBUG=false.
FLASK_DEBUG = os.environ.get("FLASK_DEBUG", "true").strip().lower() == "true"
app.debug = FLASK_DEBUG  # set explicitly (and before the guard below) so the check is meaningful

# Start the auto-checkout background job exactly once, no matter how this
# module ends up running:
#   - `python app.py` with FLASK_DEBUG on: Werkzeug's reloader re-executes this
#     whole script in a child process (WERKZEUG_RUN_MAIN=true there) after first
#     running it once in a "monitor" process that just watches files. Starting
#     the scheduler unconditionally would run two copies of the same job, so
#     only the child process starts it.
#   - `python app.py` with FLASK_DEBUG off, or Gunicorn importing `app:app` in
#     production: there's no reloader involved, so start immediately. This
#     module-level code runs once per OS process that imports it — Gunicorn
#     MUST run with a single worker (see Dockerfile), since each worker process
#     would otherwise import this module separately and start its own copy of
#     the job.
if not FLASK_DEBUG or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
    scheduler = BackgroundScheduler()
    scheduler.add_job(auto_checkout_job, "interval", minutes=2, id="auto_checkout", replace_existing=True)
    scheduler.add_job(retire_expired_accounts, "interval", hours=24, id="retire_accounts", replace_existing=True)
    scheduler.start()


if __name__ == "__main__":
    app.run(debug=FLASK_DEBUG)
